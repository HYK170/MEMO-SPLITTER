"""Manual integration test helper for MEMO SPLITTER."""

from __future__ import annotations

import sys
import tempfile
import zipfile
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage

from src.filename_builder import parse_title
from src.multimedia_copier import parse_saved_file_names
from src.splitter import SplitConfig, build_memo_output_root, split_workbook


def _tiny_png_bytes() -> bytes:
    from io import BytesIO

    buffer = BytesIO()
    PILImage.new("RGB", (1, 1), (255, 0, 0)).save(buffer, format="PNG")
    return buffer.getvalue()


TINY_PNG = _tiny_png_bytes()


def test_parse_title_first_line_only() -> None:
    title = parse_title("제목 : 회의록\n본문 내용이 길게 이어짐")
    assert title == "회의록", f"expected '회의록', got {title!r}"


def test_parse_saved_file_names() -> None:
    paths = parse_saved_file_names("a/photo.jpg\nb/note.txt\n\n c/doc.pdf ")
    assert paths == ["a/photo.jpg", "b/note.txt", "c/doc.pdf"]


def test_build_memo_output_root() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        input_path = root / "memo.xlsx"
        input_path.write_bytes(b"")
        from datetime import datetime

        first = build_memo_output_root(input_path, now=datetime(2026, 7, 14, 13, 24, 0))
        assert first.name == "memo_20260714132400"
        first.mkdir()
        second = build_memo_output_root(input_path, now=datetime(2026, 7, 14, 13, 24, 0))
        assert second.name == "memo_20260714132400_2"


def create_sample_workbook(path: Path, attach_header: str = "첨부 파일") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["App", "본문", "저장된 파일 이름", attach_header]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    rows = [
        ("Kakao", "제목 : 회의록\n본문", "images/shot.png\ndocs/memo.txt", ""),
        ("Line", "본문만 있음", "", ""),
        ("Slack", "제목 : 긴급 공지\n상세", "images/shot.png", ""),
    ]
    for idx, (app, body, saved, attach) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=app)
        ws.cell(row=idx, column=2, value=body)
        ws.cell(row=idx, column=3, value=saved)
        ws.cell(row=idx, column=4, value=attach)

    wb.save(path)
    wb.close()


def run_split(input_path: Path, multimedia: Path) -> object:
    return split_workbook(
        SplitConfig(
            input_path=input_path,
            multimedia_root=multimedia,
            sheet_name="Sheet1",
            header_row=1,
        ),
        on_log=print,
    )


def main() -> None:
    test_parse_title_first_line_only()
    test_parse_saved_file_names()
    test_build_memo_output_root()

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        input_path = root / "memo.xlsx"
        multimedia = root / "Multimedia"
        (multimedia / "images").mkdir(parents=True)
        (multimedia / "docs").mkdir(parents=True)
        (multimedia / "images" / "shot.png").write_bytes(TINY_PNG)
        (multimedia / "docs" / "memo.txt").write_text("hello", encoding="utf-8")

        create_sample_workbook(input_path)

        result = run_split(input_path, multimedia)

        print("RESULT", result)
        assert result.output_root is not None
        assert result.output_root.parent == root
        assert result.output_root.name.startswith("memo_")
        assert result.folders_created == 3
        assert result.attachments_copied == 3
        assert result.images_embedded == 2
        assert not result.attachment_skips
        assert not result.row_errors

        folder1 = result.output_root / "memo_001"
        attachments1 = folder1 / "memo_001_attach"
        assert (attachments1 / "shot.png").is_file()
        assert (attachments1 / "memo.txt").is_file()
        assert not (folder1 / "shot.png").exists()

        first_xlsx = folder1 / "memo_001_회의록.xlsx"
        assert first_xlsx.is_file()
        check_wb = load_workbook(first_xlsx)
        check_ws = check_wb.active
        assert check_ws.cell(1, 1).value == "App"
        assert check_ws.cell(2, 1).value == "Kakao"
        assert check_ws.cell(3, 1).value is None
        assert len(getattr(check_ws, "_images", [])) == 1
        check_wb.close()

        with zipfile.ZipFile(first_xlsx) as z:
            assert any(n.startswith("xl/media/") for n in z.namelist())

        for folder in sorted(result.output_root.iterdir()):
            children = []
            for p in folder.iterdir():
                if p.is_dir():
                    children.append(f"{p.name}/[{', '.join(c.name for c in p.iterdir())}]")
                else:
                    children.append(p.name)
            print("FOLDER", folder.name, children)

    # NBSP / 공백 없는 헤더도 매칭되어야 함
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        input_path = root / "memo.xlsx"
        multimedia = root / "Multimedia"
        (multimedia / "images").mkdir(parents=True)
        (multimedia / "images" / "shot.png").write_bytes(TINY_PNG)
        create_sample_workbook(input_path, attach_header="첨부\u00a0파일")
        result = run_split(input_path, multimedia)
        assert result.images_embedded == 2, result

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        input_path = root / "memo.xlsx"
        multimedia = root / "Multimedia"
        (multimedia / "images").mkdir(parents=True)
        (multimedia / "images" / "shot.png").write_bytes(TINY_PNG)
        create_sample_workbook(input_path, attach_header="첨부파일")
        result = run_split(input_path, multimedia)
        assert result.images_embedded == 2, result

    print("ALL TESTS PASSED")


if __name__ == "__main__":
    main()
