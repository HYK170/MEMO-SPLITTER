"""XLSX 이미지 진단 스크립트.

사용법:
  python scripts/diagnose_xlsx.py "INPUT.xlsx" "시트명"
"""

from __future__ import annotations

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

from src.cell_image_loader import diagnose_image_sources
from src.hyperlink_image_loader import collect_image_paths_for_row, is_image_file
from src.hyperlink_resolver import resolve_local_path
from src.xlsx_hyperlink_index import XlsxHyperlinkIndex


def main() -> None:
    if len(sys.argv) < 3:
        print('사용법: python scripts/diagnose_xlsx.py "INPUT.xlsx" "시트명"')
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    sheet_name = sys.argv[2]

    if not xlsx_path.is_file():
        print(f"파일 없음: {xlsx_path}")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        print(f"시트 없음. 사용 가능: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    print(f"파일: {xlsx_path}")
    print(f"시트: {sheet_name}")
    print("--- 진단 ---")
    for line in diagnose_image_sources(xlsx_path, ws, wb):
        print(line)

    print("--- 이미지 하이퍼링크 ---")
    image_index = XlsxHyperlinkIndex(xlsx_path, header_row=1, ws=ws, wb=wb)
    for line in image_index.summarize():
        print(line)

    resolved_images = 0
    for target in image_index.unique_targets():
        local = resolve_local_path(target, xlsx_path.parent)
        if local and local.is_file() and is_image_file(local):
            resolved_images += 1
            print(f"  로드 가능: {local}")
    print(f"--- 하이퍼링크→이미지 파일 로드 가능: {resolved_images}개 ---")

    # 샘플 행 미리보기
    for row in sorted(image_index._image_targets_by_row.keys())[:5]:
        paths, skips = collect_image_paths_for_row(image_index, row, xlsx_path.parent)
        print(f"  행 {row}: 로드 후보 {len(paths)}개")
        for skip in skips[:2]:
            print(f"    {skip}")

    wb.close()


if __name__ == "__main__":
    main()
