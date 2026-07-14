from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT_HEADER_ROW = 1
OUTPUT_DATA_ROW = 2


def list_sheet_names(xlsx_path: Path) -> list[str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def is_row_empty(ws: Worksheet, row: int, min_col: int, max_col: int) -> bool:
    for col in range(min_col, max_col + 1):
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            return False
    return True


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def copy_row(
    source_ws: Worksheet,
    target_ws: Worksheet,
    source_row: int,
    target_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for col in range(min_col, max_col + 1):
        src_cell = source_ws.cell(row=source_row, column=col)
        if isinstance(src_cell, MergedCell):
            continue
        dst_cell = target_ws.cell(row=target_row, column=col)
        dst_cell.value = src_cell.value
        copy_cell_style(src_cell, dst_cell)
        dst_cell.hyperlink = None


def copy_column_dimensions(source_ws: Worksheet, target_ws: Worksheet, min_col: int, max_col: int) -> None:
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        source_dim = source_ws.column_dimensions.get(letter)
        if source_dim is not None and source_dim.width is not None:
            target_ws.column_dimensions[letter].width = source_dim.width


def copy_merged_cells_for_split(
    source_ws: Worksheet,
    target_ws: Worksheet,
    header_row: int,
    data_row: int,
) -> None:
    row_map = {header_row: OUTPUT_HEADER_ROW, data_row: OUTPUT_DATA_ROW}
    for merged_range in source_ws.merged_cells.ranges:
        if merged_range.min_row not in row_map or merged_range.max_row not in row_map:
            continue
        if row_map[merged_range.min_row] != row_map[merged_range.max_row]:
            continue
        min_row = row_map[merged_range.min_row]
        max_row = row_map[merged_range.max_row]
        start = f"{get_column_letter(merged_range.min_col)}{min_row}"
        end = f"{get_column_letter(merged_range.max_col)}{max_row}"
        target_ws.merge_cells(f"{start}:{end}")


def create_split_workbook(
    source_ws: Worksheet,
    header_row: int,
    data_row: int,
    sheet_title: str | None = None,
) -> Workbook:
    """HEADER + 데이터 1행만 복사. 항상 1행(헤더) / 2행(데이터)."""
    wb = Workbook()
    target_ws = wb.active
    target_ws.title = sheet_title or source_ws.title

    min_col = source_ws.min_column or 1
    max_col = source_ws.max_column or 1

    copy_row(source_ws, target_ws, header_row, OUTPUT_HEADER_ROW, min_col, max_col)
    copy_row(source_ws, target_ws, data_row, OUTPUT_DATA_ROW, min_col, max_col)
    copy_column_dimensions(source_ws, target_ws, min_col, max_col)

    for source_row, target_row in (
        (header_row, OUTPUT_HEADER_ROW),
        (data_row, OUTPUT_DATA_ROW),
    ):
        source_dim = source_ws.row_dimensions.get(source_row)
        if source_dim is not None and source_dim.height is not None:
            target_ws.row_dimensions[target_row].height = source_dim.height

    copy_merged_cells_for_split(source_ws, target_ws, header_row, data_row)
    return wb
