from __future__ import annotations

import gc
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from src.filename_builder import (
    build_attachments_folder_name,
    build_row_folder_name,
    build_xlsx_filename,
)
from src.image_embedder import embed_images_in_column, ensure_pillow
from src.multimedia_copier import SAVED_NAME_COLUMN, copy_multimedia_for_row
from src.sheet_copier import create_split_workbook, is_row_empty

LogCallback = Callable[[str], None]
ProgressCallback = Callable[[int, int], None]

REQUIRED_COLUMNS = ("App", "본문", SAVED_NAME_COLUMN)
ATTACH_COLUMN = "첨부 파일"
_HEADER_WHITESPACE = re.compile(r"\s+", re.UNICODE)


@dataclass
class SplitResult:
    folders_created: int = 0
    attachments_copied: int = 0
    images_embedded: int = 0
    rows_skipped: int = 0
    attachment_skips: list[str] = field(default_factory=list)
    row_errors: list[str] = field(default_factory=list)
    output_root: Path | None = None


@dataclass
class SplitConfig:
    input_path: Path
    multimedia_root: Path
    sheet_name: str
    header_row: int


def build_memo_output_root(input_path: Path, now: datetime | None = None) -> Path:
    """INPUT XLSX와 같은 경로에 {원본파일명}_{YYYYMMDDHHMMSS} 폴더 경로를 만든다."""
    parent = input_path.resolve().parent
    base_name = input_path.stem
    stamp = (now or datetime.now()).strftime("%Y%m%d%H%M%S")
    candidate = parent / f"{base_name}_{stamp}"
    if not candidate.exists():
        return candidate
    counter = 2
    while True:
        numbered = parent / f"{base_name}_{stamp}_{counter}"
        if not numbered.exists():
            return numbered
        counter += 1


def validate_config(config: SplitConfig) -> None:
    if not config.input_path.is_file():
        raise FileNotFoundError(f"INPUT XLSX를 찾을 수 없습니다: {config.input_path}")
    if config.input_path.suffix.lower() != ".xlsx":
        raise ValueError("INPUT 파일은 .xlsx 확장자여야 합니다.")
    if not config.multimedia_root.is_dir():
        raise FileNotFoundError(f"Multimedia 폴더를 찾을 수 없습니다: {config.multimedia_root}")
    if config.header_row < 1:
        raise ValueError("HEADER ROW는 1 이상이어야 합니다.")


def split_workbook(
    config: SplitConfig,
    on_log: LogCallback | None = None,
    on_progress: ProgressCallback | None = None,
) -> SplitResult:
    validate_config(config)

    def log(message: str) -> None:
        if on_log:
            on_log(message)

    output_root = build_memo_output_root(config.input_path)
    output_root.mkdir(parents=True, exist_ok=False)
    log(f"OUTPUT: {output_root}")

    log(f"Python: {sys.executable}")
    try:
        pillow_version = ensure_pillow()
        log(f"Pillow: {pillow_version}")
    except RuntimeError as exc:
        log(f"경고: {exc}")
        pillow_version = None

    wb = load_workbook(config.input_path, data_only=False)
    try:
        if config.sheet_name not in wb.sheetnames:
            raise ValueError(f"시트를 찾을 수 없습니다: {config.sheet_name}")

        ws = wb[config.sheet_name]
        if config.header_row > (ws.max_row or 0):
            raise ValueError("HEADER ROW가 시트 범위를 벗어났습니다.")

        column_map = _build_column_map(ws, config.header_row)
        _ensure_required_columns(column_map)
        attach_col = _find_column(column_map, ATTACH_COLUMN)
        if attach_col is None:
            headers = ", ".join(column_map.keys()) or "(없음)"
            log(
                f"경고: '{ATTACH_COLUMN}' 열을 찾지 못했습니다. "
                f"이미지 임베드를 건너뜁니다. 현재 헤더: {headers}"
            )
        elif pillow_version is None:
            log(f"경고: Pillow 없음 — '{ATTACH_COLUMN}' 열 이미지 임베드를 건너뜁니다.")
        else:
            log(f"이미지 임베드 대상 열: '{ATTACH_COLUMN}' (열 {attach_col}), 행 2")

        base_name = config.input_path.stem
        min_col = ws.min_column or 1
        max_col = ws.max_column or 1
        first_data_row = config.header_row + 1
        last_data_row = ws.max_row or config.header_row

        data_rows = [
            row
            for row in range(first_data_row, last_data_row + 1)
            if not is_row_empty(ws, row, min_col, max_col)
        ]
        result = SplitResult(output_root=output_root)
        result.rows_skipped = (last_data_row - first_data_row + 1) - len(data_rows)
        total = len(data_rows)
        row_index = 0

        for data_row in data_rows:
            row_index += 1
            if on_progress:
                on_progress(row_index, total)

            folder_name = build_row_folder_name(base_name, row_index)
            row_folder = output_root / folder_name

            try:
                row_folder.mkdir(parents=True, exist_ok=True)
            except OSError as exc:
                message = f"행 {data_row}: 폴더 생성 실패 ({row_folder}) - {exc}"
                result.row_errors.append(message)
                log(message)
                continue

            body_value = ws.cell(row=data_row, column=column_map["본문"]).value
            saved_names = ws.cell(row=data_row, column=column_map[SAVED_NAME_COLUMN]).value
            xlsx_name = build_xlsx_filename(base_name, row_index, body_value)
            xlsx_path = row_folder / xlsx_name

            attachments_folder = row_folder / build_attachments_folder_name(base_name, row_index)
            copy_result = copy_multimedia_for_row(
                saved_names,
                config.multimedia_root,
                attachments_folder,
            )
            result.attachments_copied += len(copy_result.copied)
            result.attachment_skips.extend(copy_result.skipped)

            try:
                out_wb = create_split_workbook(ws, config.header_row, data_row, config.sheet_name)
                out_ws = out_wb.active
                embedded = 0
                embed_failures: list[str] = []
                if attach_col and pillow_version and copy_result.image_paths:
                    embedded, embed_failures = embed_images_in_column(
                        out_ws,
                        copy_result.image_paths,
                        attach_col,
                    )
                    result.images_embedded += embedded
                out_wb.save(xlsx_path)
                out_wb.close()
                del out_wb
            except Exception as exc:
                message = f"행 {data_row}: XLSX 저장 실패 - {exc}"
                result.row_errors.append(message)
                log(message)
                continue

            result.folders_created += 1
            log(f"행 {data_row}: {row_folder.name} 생성 ({xlsx_name})")
            for name in copy_result.copied:
                log(f"  첨부 복사: {name}")
            if embedded:
                log(f"  이미지 임베드: {embedded}개 -> {ATTACH_COLUMN} 열{attach_col}/행2")
            elif copy_result.image_paths and attach_col is None:
                log(f"  이미지 임베드 스킵: '{ATTACH_COLUMN}' 열 없음")
            elif copy_result.image_paths and pillow_version is None:
                log("  이미지 임베드 스킵: Pillow 미설치")
            elif copy_result.copied and not copy_result.image_paths:
                log("  이미지 임베드 스킵: 복사된 첨부 중 이미지 확장자 없음")
            for failure in embed_failures:
                log(f"  이미지 임베드 실패: {failure}")
            for skipped in copy_result.skipped:
                log(f"  첨부 스킵: {skipped}")

            if row_index % 100 == 0:
                gc.collect()

        return result
    finally:
        wb.close()
        gc.collect()


def _normalize_header(value: object) -> str:
    """헤더 문자열의 공백(NBSP 포함)을 정규화한다."""
    return _HEADER_WHITESPACE.sub(" ", str(value).strip())


def _build_column_map(ws, header_row: int) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for col in range(ws.min_column or 1, (ws.max_column or 1) + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        column_map[_normalize_header(value)] = col
    return column_map


def _find_column(column_map: dict[str, int], name: str) -> int | None:
    """정확한 이름 또는 공백 제거 후 이름으로 열을 찾는다."""
    normalized = _normalize_header(name)
    if normalized in column_map:
        return column_map[normalized]
    compact = normalized.replace(" ", "")
    for key, col in column_map.items():
        if key.replace(" ", "") == compact:
            return col
    return None


def _ensure_required_columns(column_map: dict[str, int]) -> None:
    missing = [name for name in REQUIRED_COLUMNS if _find_column(column_map, name) is None]
    if missing:
        raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing)}")
    # 필수 컬럼도 정규화/공백 제거 매칭으로 통일된 키를 쓰도록 보정
    for name in REQUIRED_COLUMNS:
        col = _find_column(column_map, name)
        if col is not None:
            column_map[name] = col
