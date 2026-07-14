from __future__ import annotations

import math
import sys
from io import BytesIO
from pathlib import Path

from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import (
    EMU_to_pixels,
    cm_to_EMU,
    pixels_to_EMU,
    pixels_to_points,
)
from openpyxl.worksheet.worksheet import Worksheet

from src.sheet_copier import OUTPUT_DATA_ROW

# 고정 표시 박스 (이미지 비율 유지, 남는 영역은 검정)
MAX_WIDTH_CM = 2.54
MAX_HEIGHT_CM = 2.36
GAP_CM = 0.1  # 여러 장 세로 쌓을 때 간격
ROW_PADDING_CM = 0.15  # 행 높이 여유
# 캔버스 해상도 (표시 cm는 유지, 비트맵만 더 선명하게)
RENDER_DPI = 150


def ensure_pillow() -> str:
    """openpyxl Image가 Pillow를 쓰도록 보장한다. 성공 시 버전 문자열 반환."""
    try:
        from PIL import Image as PILImage
    except ImportError as exc:
        raise RuntimeError(
            "이미지 임베드에 Pillow가 필요합니다. "
            f"현재 Python: {sys.executable}\n"
            '설치: python -m pip install "Pillow>=10.0.0"'
        ) from exc

    # openpyxl이 Pillow 없이 먼저 import되면 PILImage=False로 고정된다.
    import openpyxl.drawing.image as oxl_image

    if not oxl_image.PILImage:
        oxl_image.PILImage = PILImage

    version = getattr(PILImage, "__version__", None)
    if version is None:
        import PIL

        version = getattr(PIL, "__version__", "unknown")
    return str(version)


def _cm_to_pixels(cm: float) -> float:
    return float(EMU_to_pixels(cm_to_EMU(cm)))


def _display_box_pixels() -> tuple[int, int]:
    """엑셀에 표시할 고정 박스 크기(픽셀, openpyxl 기준)."""
    return (
        max(1, int(round(_cm_to_pixels(MAX_WIDTH_CM)))),
        max(1, int(round(_cm_to_pixels(MAX_HEIGHT_CM)))),
    )


def _render_box_pixels() -> tuple[int, int]:
    """검정 배경 합성용 캔버스 픽셀 크기."""
    return (
        max(1, int(round(MAX_WIDTH_CM * RENDER_DPI / 2.54))),
        max(1, int(round(MAX_HEIGHT_CM * RENDER_DPI / 2.54))),
    )


def _text_display_units(text: str) -> float:
    """엑셀 열 너비 단위에 가깝게 문자 폭을 계산한다(한글≈2)."""
    width = 0.0
    for ch in text:
        width += 2.0 if ord(ch) > 0x2E80 else 1.0
    return width


def _estimate_wrapped_row_height(ws: Worksheet, row: int) -> float:
    """자동 줄바꿈이 필요할 때 대략적인 행 높이(pt)."""
    max_height = 0.0
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value in (None, ""):
            continue

        font_size = 11.0
        if cell.font is not None and cell.font.size:
            font_size = float(cell.font.size)

        wrap = True
        if cell.alignment is not None and cell.alignment.wrap_text is False:
            wrap = False

        text = str(cell.value).replace("\r\n", "\n").replace("\r", "\n")
        if not wrap:
            lines = text.count("\n") + 1
        else:
            letter = get_column_letter(col)
            dim = ws.column_dimensions.get(letter)
            col_width = float(dim.width) if dim is not None and dim.width else 8.43
            chars_per_line = max(1.0, col_width)
            lines = 0
            for part in text.split("\n"):
                units = _text_display_units(part) if part else 1.0
                lines += max(1, int(math.ceil(units / chars_per_line)))

        max_height = max(max_height, lines * font_size * 1.3)

    return min(409.0, max_height)


def _compose_boxed_png(path: Path) -> bytes:
    """비율 유지로 맞춘 이미지를 고정 박스 중앙에 두고 나머지는 검정으로 채운 PNG."""
    from PIL import Image as PILImage

    box_w, box_h = _render_box_pixels()
    with PILImage.open(path) as src:
        rgba = src.convert("RGBA")
        # 박스 안에 맞게 축소만 하고, 작은 이미지는 확대하지 않음
        scale = min(box_w / rgba.width, box_h / rgba.height, 1.0)
        fitted_w = max(1, int(rgba.width * scale))
        fitted_h = max(1, int(rgba.height * scale))
        fitted = rgba.resize((fitted_w, fitted_h), PILImage.Resampling.LANCZOS)

    canvas = PILImage.new("RGB", (box_w, box_h), (0, 0, 0))
    offset = ((box_w - fitted_w) // 2, (box_h - fitted_h) // 2)
    canvas.paste(fitted, offset, fitted)

    buffer = BytesIO()
    canvas.save(buffer, format="PNG")
    return buffer.getvalue()


def embed_images_in_column(
    ws: Worksheet,
    image_paths: list[Path],
    column: int,
    row: int = OUTPUT_DATA_ROW,
) -> tuple[int, list[str]]:
    """이미지 파일을 지정 열 셀에 임베드한다.

    각 이미지는 항상 2.54x2.36cm 박스로 들어가고, 비율 유지 + 검정 레터박스.
    여러 장이면 세로로 쌓고, 2행 높이는 필요할 때만 늘린다(줄이지 않음).

    Returns:
        (성공 개수, 실패 메시지 목록)
    """
    if column < 1 or not image_paths:
        return 0, []

    try:
        ensure_pillow()
    except RuntimeError as exc:
        return 0, [str(exc)]

    display_w, display_h = _display_box_pixels()
    gap_px = _cm_to_pixels(GAP_CM)
    added = 0
    failures: list[str] = []
    cursor_y = 0.0

    for path in image_paths:
        try:
            png_bytes = _compose_boxed_png(path)
            image = Image(BytesIO(png_bytes))
            image.width = display_w
            image.height = display_h

            marker = AnchorMarker(
                col=column - 1,
                row=row - 1,
                colOff=0,
                rowOff=int(pixels_to_EMU(cursor_y)),
            )
            image.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(
                    int(pixels_to_EMU(display_w)),
                    int(pixels_to_EMU(display_h)),
                ),
            )
            image._cached_bytes = png_bytes  # noqa: SLF001
            image._data = lambda data=png_bytes: data  # noqa: SLF001
            ws.add_image(image)

            cursor_y += display_h + gap_px
            added += 1
        except Exception as exc:
            failures.append(f"{path.name}: 삽입 실패 ({exc})")

    if added:
        total_height_px = added * display_h + (added - 1) * gap_px
        image_needed = pixels_to_points(total_height_px + _cm_to_pixels(ROW_PADDING_CM))
        image_needed = min(409.0, max(15.0, float(image_needed)))
        current = ws.row_dimensions[row].height
        wrap_needed = _estimate_wrapped_row_height(ws, row)
        # 이미지 / 자동줄바꿈 / 기존 명시 높이 중 최대값. 절대 낮추지 않음.
        target = max(image_needed, wrap_needed)
        if current is not None:
            target = max(target, float(current))
        if current is None or float(current) < target:
            ws.row_dimensions[row].height = target

    return added, failures
