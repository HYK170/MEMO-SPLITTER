from __future__ import annotations

import io
import zipfile
from copy import copy
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree as ET

from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

HYPERLINK_TAGS = {
    "{http://schemas.openxmlformats.org/drawingml/2006/main}hlinkClick",
    "{http://schemas.openxmlformats.org/drawingml/2006/main}hlinkHover",
}

TINY_PNG_BYTES = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc``\x00\x00"
    b"\x00\x02\x00\x01\xe2!\xbc3\x00\x00\x00\x00IEND\xaeB`\x82"
)


def get_image_row_range(image: Image) -> tuple[int, int] | None:
    anchor = getattr(image, "anchor", None)
    if anchor is None:
        return None

    if isinstance(anchor, str):
        parsed = _parse_string_anchor(anchor)
        if parsed is None:
            return None
        from_row, to_row, _ = parsed
        return from_row, to_row

    from_anchor = getattr(anchor, "_from", None)
    if from_anchor is None:
        return None

    from_row = int(from_anchor.row) + 1
    to_anchor = getattr(anchor, "to", None) or getattr(anchor, "_to", None)
    if to_anchor is not None and getattr(to_anchor, "row", None) is not None:
        to_row = int(to_anchor.row) + 1
    else:
        to_row = from_row

    if to_row < from_row:
        to_row = from_row
    return from_row, to_row


def _parse_string_anchor(anchor: str) -> tuple[int, int, int] | None:
    cleaned = anchor.replace("$", "").strip()
    try:
        _, row = coordinate_from_string(cleaned)
    except ValueError:
        return None
    return row, row, 0


def get_assigned_rows(from_row: int, to_row: int, header_row: int = 1) -> range:
    if from_row == to_row:
        first_data_row = header_row + 1
        if from_row - 1 == first_data_row:
            return range(from_row - 1, from_row + 1)
        return range(from_row, from_row + 2)
    return range(from_row, to_row + 1)


def image_matches_row(image: Image, data_row: int, header_row: int = 1) -> bool:
    row_range = get_image_row_range(image)
    if row_range is None:
        return False
    from_row, to_row = row_range
    return data_row in get_assigned_rows(from_row, to_row, header_row)


def get_image_row(image: Image) -> int | None:
    row_range = get_image_row_range(image)
    if row_range is None:
        return None
    return row_range[0]


def index_images_by_row(ws: Worksheet, header_row: int = 1, max_row: int | None = None) -> dict[int, list[Image]]:
    images_by_row: dict[int, list[Image]] = {}
    if max_row is None:
        max_row = ws.max_row or 1

    for image in getattr(ws, "_images", []):
        _read_image_bytes(image)
        row_range = get_image_row_range(image)
        if row_range is None:
            continue
        from_row, to_row = row_range
        for row in get_assigned_rows(from_row, to_row, header_row):
            if row > max_row:
                continue
            images_by_row.setdefault(row, []).append(image)
    return images_by_row


def emu_to_pixels(emu: int) -> int:
    return max(1, int(round(emu / 9525)))


def guess_media_format(data: bytes, media_path: str = "") -> str:
    ext = PurePosixPath(media_path.replace("\\", "/")).suffix.lower().lstrip(".")
    if ext in {"jpg", "jpeg", "png", "gif", "bmp", "emf", "wmf", "tiff", "tif", "webp"}:
        return "jpeg" if ext == "jpg" else ext
    if data.startswith(b"\x89PNG"):
        return "png"
    if data.startswith(b"\xff\xd8"):
        return "jpeg"
    if data.startswith(b"GIF8"):
        return "gif"
    if len(data) > 44 and data[40:44] == b" EMF":
        return "emf"
    if data.startswith(b"\xd7\xcd\xc6\x9a") or data.startswith(b"\x01\x00\t\x00"):
        return "wmf"
    return "png"


def create_openpyxl_image_from_bytes(
    data: bytes,
    *,
    width: int | None = None,
    height: int | None = None,
    media_path: str = "",
    source_format: str | None = None,
) -> Image | None:
    try:
        image = Image(io.BytesIO(data))
    except Exception:
        try:
            image = Image(io.BytesIO(TINY_PNG_BYTES))
        except Exception:
            return None
        image.format = source_format or guess_media_format(data, media_path)
        image.width = width or 200
        image.height = height or 150

    if width:
        image.width = width
    if height:
        image.height = height
    if source_format:
        image.format = source_format
    elif media_path:
        image.format = guess_media_format(data, media_path)

    image._cached_bytes = data  # noqa: SLF001
    image._data = lambda: data  # noqa: SLF001
    return image


def can_create_openpyxl_image(data: bytes, media_path: str = "") -> bool:
    return create_openpyxl_image_from_bytes(data, media_path=media_path) is not None


def strip_image_hyperlinks(image: Image) -> None:
    anchor = getattr(image, "anchor", None)
    if anchor is None:
        return

    pic = getattr(anchor, "pic", None)
    if pic is not None:
        nv_pic_pr = getattr(pic, "nvPicPr", None)
        if nv_pic_pr is not None:
            c_nv_pr = getattr(nv_pic_pr, "cNvPr", None)
            if c_nv_pr is not None:
                if hasattr(c_nv_pr, "hlinkClick"):
                    c_nv_pr.hlinkClick = None
                if hasattr(c_nv_pr, "hlinkHover"):
                    c_nv_pr.hlinkHover = None


def copy_image_for_row(source_image: Image) -> Image:
    data = _read_image_bytes(source_image)
    new_image = create_openpyxl_image_from_bytes(
        data,
        width=int(source_image.width),
        height=int(source_image.height),
        source_format=getattr(source_image, "format", None) or "",
    )
    if new_image is None:
        raise ValueError("시각 객체 데이터를 복사할 수 없습니다.")
    new_image.anchor = copy(source_image.anchor)
    strip_image_hyperlinks(new_image)
    return new_image


def _read_image_bytes(source_image: Image) -> bytes:
    cached = getattr(source_image, "_cached_bytes", None)
    if cached is not None:
        return cached

    ref = getattr(source_image, "ref", None)
    if isinstance(ref, (str, Path)):
        ref_path = Path(ref)
        if ref_path.is_file():
            data_bytes = ref_path.read_bytes()
            source_image._cached_bytes = data_bytes  # noqa: SLF001
            return data_bytes

    try:
        data = source_image._data()
    except (ValueError, OSError):
        internal = getattr(source_image, "_image", None)
        if internal is None:
            raise
        internal.seek(0)
        data = internal.read()

    if isinstance(data, (bytes, bytearray)):
        data_bytes = bytes(data)
    else:
        data_bytes = bytes(data.read())

    source_image._cached_bytes = data_bytes  # noqa: SLF001
    return data_bytes


def add_row_images(target_ws: Worksheet, images: list[Image]) -> None:
    seen_ids: set[int] = set()
    for image in images:
        image_id = id(image)
        if image_id in seen_ids:
            continue
        seen_ids.add(image_id)
        copied = copy_image_for_row(image)
        target_ws.add_image(copied)


def get_image_cell_coordinate(image: Image) -> str | None:
    row_range = get_image_row_range(image)
    anchor = getattr(image, "anchor", None)
    if row_range is None or anchor is None:
        return None
    from_anchor = getattr(anchor, "_from", None)
    if from_anchor is None:
        return None
    col = int(from_anchor.col) + 1
    return f"{get_column_letter(col)}{row_range[0]}"


def strip_hyperlinks_from_xlsx_file(xlsx_path) -> None:
    buffer = io.BytesIO()
    removed_rel_ids: set[str] = set()

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/drawings/") and item.filename.endswith(".xml"):
                    data, rel_ids = _strip_drawing_hyperlinks(data)
                    removed_rel_ids.update(rel_ids)
                elif item.filename.startswith("xl/drawings/_rels/") and item.filename.endswith(".rels"):
                    if removed_rel_ids:
                        data = _remove_relationships(data, removed_rel_ids)
                zout.writestr(item, data)

    with open(xlsx_path, "wb") as out_file:
        out_file.write(buffer.getvalue())


def _strip_drawing_hyperlinks(data: bytes) -> tuple[bytes, set[str]]:
    removed_rel_ids: set[str] = set()
    root = ET.fromstring(data)
    for tag in HYPERLINK_TAGS:
        for element in root.iter(tag):
            rel_id = element.attrib.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            if rel_id:
                removed_rel_ids.add(rel_id)
            parent = _find_parent(root, element)
            if parent is not None:
                parent.remove(element)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True), removed_rel_ids


def _find_parent(root: ET.Element, child: ET.Element) -> ET.Element | None:
    for parent in root.iter():
        if child in list(parent):
            return parent
    return None


def _remove_relationships(data: bytes, rel_ids: set[str]) -> bytes:
    root = ET.fromstring(data)
    rel_tag = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    for rel in list(root.findall(rel_tag)):
        if rel.attrib.get("Id") in rel_ids:
            root.remove(rel)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)
