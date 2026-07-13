from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

REL_ID_ATTR = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
REL_EMBED_ATTR = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
DISPIMG_PATTERN = re.compile(r'DISPIMG\s*\(\s*"([^"]+)"', re.IGNORECASE)
DIAGNOSE_VERSION = "2026-07-13-e"


def load_cell_images(xlsx_path: Path, ws: Worksheet, wb=None) -> dict[int, list[Image]]:
    images_by_row: dict[int, list[Image]] = {}

    with zipfile.ZipFile(xlsx_path, "r") as archive:
        sheet_path = _resolve_sheet_path(archive, ws, wb)
        if not sheet_path:
            return {}

        id_to_media = _build_cellimage_id_map(archive)
        if id_to_media:
            _load_dispimg_cells_from_sheet_xml(archive, sheet_path, id_to_media, images_by_row)
            _load_dispimg_cells_from_openpyxl(ws, id_to_media, xlsx_path, images_by_row)

        _load_richdata_cell_images(archive, sheet_path, images_by_row)

    return images_by_row


def diagnose_image_sources(xlsx_path: Path, ws: Worksheet, wb=None) -> list[str]:
    lines: list[str] = []
    lines.append(f"진단 버전: {DIAGNOSE_VERSION}")
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        names = archive.namelist()
        xml_sheet_names = list_workbook_sheet_names(archive)
        lines.append(f"workbook.xml 시트명: {xml_sheet_names}")
        lines.append(f"openpyxl 시트명: {ws.title!r} / ws.path: {getattr(ws, 'path', None)!r}")

        sheet_path = _resolve_sheet_path(archive, ws, wb)
        lines.append(f"시트 XML: {sheet_path or '찾을 수 없음'}")

        from src.sheet_path import get_effective_max_row

        xml_max_row = get_effective_max_row(ws, archive, sheet_path) if sheet_path else (ws.max_row or 1)
        lines.append(f"openpyxl max_row: {ws.max_row or 1} / sheet XML max_row: {xml_max_row}")

        if sheet_path:
            from src.sheet_path import list_sheet_relationships

            rels = list_sheet_relationships(archive, sheet_path)
            lines.append(f"시트 rels: {rels or '없음'}")

        from src.sheet_path import discover_drawing_paths, discover_vml_paths, list_package_drawing_parts

        drawing_paths = discover_drawing_paths(archive, sheet_path, ws, wb) if sheet_path else []
        vml_paths = discover_vml_paths(archive, sheet_path, ws, wb) if sheet_path else []
        lines.append(f"drawing XML: {len(drawing_paths)}개 {drawing_paths}")
        lines.append(f"VML drawing: {len(vml_paths)}개 {vml_paths}")

        package_drawings = list_package_drawing_parts(archive)
        lines.append(f"패키지 drawing 전체: {len(package_drawings)}개")
        for drawing_path in package_drawings[:5]:
            root = ET.fromstring(archive.read(drawing_path))
            blip_count = len(_iter_local_name(root, "blip"))
            anchor_count = (
                len(_iter_local_name(root, "oneCellAnchor"))
                + len(_iter_local_name(root, "twoCellAnchor"))
                + len(_iter_local_name(root, "absoluteAnchor"))
            )
            lines.append(f"  {drawing_path}: anchor {anchor_count}개, blip {blip_count}개")
        if len(package_drawings) > 5:
            lines.append(f"  ... 외 {len(package_drawings) - 5}개")

        cellimage_parts = [n for n in names if "cellimage" in n.lower() and n.endswith(".xml") and "_rels" not in n]
        lines.append(f"cellimages.xml: {cellimage_parts or '없음'}")

        media_files = [n for n in names if n.startswith("xl/media/")]
        lines.append(f"xl/media 파일: {len(media_files)}개")

        id_map = _build_cellimage_id_map(archive)
        lines.append(f"cellimages ID 매핑: {len(id_map)}개")
        if id_map:
            lines.append(f"  예시 ID: {next(iter(id_map.keys()))}")

        if sheet_path:
            dispimg_ids = _find_dispimg_ids_in_sheet(archive, sheet_path)
            lines.append(f"DISPIMG 셀: {len(dispimg_ids)}개 {dispimg_ids[:3]}")

        rich_rows = _count_richdata_cells(archive, sheet_path) if sheet_path else 0
        lines.append(f"richData(vm) 셀: {rich_rows}개")

        if sheet_path and drawing_paths:
            from src.drawing_image_loader import count_drawing_load_stats

            anchor_count = 0
            blip_count = 0
            resolved_count = 0
            assigned_at_openpyxl_max = 0
            assigned_at_xml_max = 0
            anchor_row_min: int | None = None
            anchor_row_max: int | None = None
            for drawing_path in drawing_paths:
                drawing_rels = _get_rels_map(archive, drawing_path)
                rows, blips, resolved, assigned, row_min, row_max = count_drawing_load_stats(
                    archive,
                    drawing_path,
                    drawing_rels,
                    max_row=ws.max_row or 1,
                )
                anchor_count += rows
                blip_count += blips
                resolved_count += resolved
                assigned_at_openpyxl_max += assigned
                _, _, _, assigned_xml, _, _ = count_drawing_load_stats(
                    archive,
                    drawing_path,
                    drawing_rels,
                    max_row=xml_max_row,
                )
                assigned_at_xml_max += assigned_xml
                if row_min is not None:
                    anchor_row_min = row_min if anchor_row_min is None else min(anchor_row_min, row_min)
                if row_max is not None:
                    anchor_row_max = row_max if anchor_row_max is None else max(anchor_row_max, row_max)
            lines.append(
                f"drawing 로드 가능: anchor {anchor_count}개, blip {blip_count}개, "
                f"media 해석 {resolved_count}개"
            )
            if anchor_row_min is not None and anchor_row_max is not None:
                lines.append(f"drawing anchor 행 범위: {anchor_row_min}~{anchor_row_max}")
            lines.append(
                f"행 배정 가능: openpyxl max_row({ws.max_row or 1})={assigned_at_openpyxl_max}개 / "
                f"sheet XML max_row({xml_max_row})={assigned_at_xml_max}개"
            )

        openpyxl_images = getattr(ws, "_images", [])
        lines.append(f"openpyxl _images: {len(openpyxl_images)}개")
        for idx, image in enumerate(openpyxl_images, start=1):
            from src.image_handler import get_image_row_range

            row_range = get_image_row_range(image)
            lines.append(f"  image[{idx}] anchor row: {row_range}")

    return lines


def _resolve_sheet_path(archive: zipfile.ZipFile, ws: Worksheet, wb=None) -> str | None:
    from src.sheet_path import resolve_sheet_path

    return resolve_sheet_path(archive, ws, wb)


def _list_workbook_sheet_names(archive: zipfile.ZipFile) -> list[str]:
    from src.sheet_path import list_workbook_sheet_names

    return list_workbook_sheet_names(archive)


list_workbook_sheet_names = _list_workbook_sheet_names


def _find_sheet_related_paths(archive: zipfile.ZipFile, sheet_path: str, keyword: str) -> list[str]:
    from src.sheet_path import find_sheet_related_paths

    return find_sheet_related_paths(archive, sheet_path, keyword)


def _get_rels_map(archive: zipfile.ZipFile, resource_path: str) -> dict[str, str]:
    from src.sheet_path import get_rels_map

    return get_rels_map(archive, resource_path)


def _find_cellimages_part(archive: zipfile.ZipFile) -> str | None:
    for name in archive.namelist():
        lower = name.lower()
        if lower.endswith("cellimages.xml") and "_rels" not in lower:
            return name
    return None


def _build_cellimage_id_map(archive: zipfile.ZipFile) -> dict[str, str]:
    cellimages_path = _find_cellimages_part(archive)
    if not cellimages_path:
        return {}

    rels_path = str(Path(cellimages_path).parent / "_rels" / f"{Path(cellimages_path).name}.rels")
    if rels_path not in archive.namelist():
        return {}

    embed_to_media = _get_rels_map(archive, cellimages_path)
    root = ET.fromstring(archive.read(cellimages_path))

    id_map: dict[str, str] = {}
    for cell_image in _iter_local_name(root, "cellImage"):
        image_id = _find_image_name(cell_image)
        media_path = _find_embedded_media(cell_image, embed_to_media)
        if image_id and media_path:
            id_map[image_id] = media_path

    for pic in _iter_local_name(root, "pic"):
        image_id = _find_image_name(pic)
        media_path = _find_embedded_media(pic, embed_to_media)
        if image_id and media_path:
            id_map[image_id] = media_path

    return id_map


def _find_image_name(element: ET.Element) -> str | None:
    for child in element.iter():
        if child.tag.endswith("}cNvPr") or child.tag == "cNvPr":
            name = child.attrib.get("name")
            if name:
                return name
    return None


def _find_embedded_media(element: ET.Element, embed_to_media: dict[str, str]) -> str | None:
    for blip in _iter_local_name(element, "blip"):
        rel_id = blip.attrib.get(REL_EMBED_ATTR) or blip.attrib.get(REL_ID_ATTR)
        if rel_id and rel_id in embed_to_media:
            return embed_to_media[rel_id]
    return None


def _load_dispimg_cells_from_sheet_xml(
    archive: zipfile.ZipFile,
    sheet_path: str,
    id_to_media: dict[str, str],
    images_by_row: dict[int, list[Image]],
) -> None:
    root = ET.fromstring(archive.read(sheet_path))
    for cell in _iter_local_name(root, "c"):
        cell_ref = cell.attrib.get("r")
        if not cell_ref:
            continue
        formula = _cell_formula_text(cell)
        if not formula:
            continue
        match = DISPIMG_PATTERN.search(formula)
        if not match:
            continue
        image_id = match.group(1)
        media_path = id_to_media.get(image_id)
        if not media_path or media_path not in archive.namelist():
            continue
        row, col = _split_cell_ref(cell_ref)
        image = _create_cell_image(archive.read(media_path), row, col)
        if image:
            images_by_row.setdefault(row, []).append(image)


def _load_dispimg_cells_from_openpyxl(
    ws: Worksheet,
    id_to_media: dict[str, str],
    xlsx_path: Path,
    images_by_row: dict[int, list[Image]],
) -> None:
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        for row in range(ws.min_row or 1, (ws.max_row or 1) + 1):
            for col in range(ws.min_column or 1, (ws.max_column or 1) + 1):
                value = ws.cell(row=row, column=col).value
                if not isinstance(value, str):
                    continue
                match = DISPIMG_PATTERN.search(value)
                if not match:
                    continue
                media_path = id_to_media.get(match.group(1))
                if not media_path or media_path not in archive.namelist():
                    continue
                image = _create_cell_image(archive.read(media_path), row, col - 1)
                if image:
                    images_by_row.setdefault(row, []).append(image)


def _find_dispimg_ids_in_sheet(archive: zipfile.ZipFile, sheet_path: str) -> list[str]:
    root = ET.fromstring(archive.read(sheet_path))
    ids: list[str] = []
    for cell in _iter_local_name(root, "c"):
        formula = _cell_formula_text(cell)
        if not formula:
            continue
        match = DISPIMG_PATTERN.search(formula)
        if match:
            ids.append(match.group(1))
    return ids


def _cell_formula_text(cell: ET.Element) -> str:
    parts: list[str] = []
    for child in cell:
        if child.tag.endswith("}f") or child.tag == "f":
            if child.text:
                parts.append(child.text)
    if parts:
        return "".join(parts)
    if cell.text and "DISPIMG" in cell.text.upper():
        return cell.text
    return ""


def _split_cell_ref(cell_ref: str) -> tuple[int, int]:
    col_text = "".join(ch for ch in cell_ref if ch.isalpha())
    row_text = "".join(ch for ch in cell_ref if ch.isdigit())
    return int(row_text), column_index_from_string(col_text) - 1


def _create_cell_image(data: bytes, row: int, col: int) -> Image | None:
    try:
        image = Image(io.BytesIO(data))
    except Exception:
        return None
    image._cached_bytes = data  # noqa: SLF001
    image._data = lambda: data  # noqa: SLF001
    marker = AnchorMarker(col=col, row=row - 1, colOff=0, rowOff=0)
    image.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(1_000_000, 1_000_000))
    return image


def _load_richdata_cell_images(
    archive: zipfile.ZipFile,
    sheet_path: str,
    images_by_row: dict[int, list[Image]],
) -> None:
    metadata_path = "xl/metadata.xml"
    if metadata_path not in archive.namelist():
        return

    rel_slots = _load_richvalue_rel_slots(archive)
    if not rel_slots:
        return

    root = ET.fromstring(archive.read(sheet_path))
    for cell in _iter_local_name(root, "c"):
        if "vm" not in cell.attrib:
            continue
        cell_ref = cell.attrib.get("r")
        if not cell_ref:
            continue
        try:
            vm_index = int(cell.attrib["vm"])
        except ValueError:
            continue
        rich_index = _resolve_rich_index(archive, vm_index)
        if rich_index is None:
            continue
        media_path = _resolve_media_from_rich_index(archive, rich_index, rel_slots)
        if not media_path or media_path not in archive.namelist():
            continue
        row, col = _split_cell_ref(cell_ref)
        image = _create_cell_image(archive.read(media_path), row, col)
        if image:
            images_by_row.setdefault(row, []).append(image)


def _load_richvalue_rel_slots(archive: zipfile.ZipFile) -> list[str]:
    for name in archive.namelist():
        if name.startswith("xl/richData/") and name.endswith("richValueRel.xml"):
            root = ET.fromstring(archive.read(name))
            rel_ids: list[str] = []
            for rel in root.iter():
                if rel.tag.endswith("}rel") or rel.tag == "rel":
                    rel_id = rel.attrib.get("r:id") or rel.attrib.get(REL_ID_ATTR)
                    if rel_id:
                        rel_ids.append(rel_id)
            if rel_ids:
                rels_map = _get_rels_map(archive, name)
                return [rels_map[r] for r in rel_ids if r in rels_map]
    return []


def _resolve_rich_index(archive: zipfile.ZipFile, vm_index: int) -> int | None:
    root = ET.fromstring(archive.read("xl/metadata.xml"))
    blocks = [el for el in root.iter() if el.tag.endswith("}bk") or el.tag == "bk"]
    if vm_index <= 0 or vm_index > len(blocks):
        return None
    block = blocks[vm_index - 1]
    for child in block.iter():
        if child.tag.endswith("}v") or child.tag == "v":
            try:
                return int(child.text or "")
            except ValueError:
                continue
    return None


def _resolve_media_from_rich_index(
    archive: zipfile.ZipFile,
    rich_index: int,
    rel_slots: list[str],
) -> str | None:
    for name in archive.namelist():
        if not name.startswith("xl/richData/") or not name.endswith("richvalue.xml"):
            continue
        root = ET.fromstring(archive.read(name))
        rvs = [el for el in root.iter() if el.tag.endswith("}rv") or el.tag == "rv"]
        if rich_index < 0 or rich_index >= len(rvs):
            continue
        rv = rvs[rich_index]
        for child in rv.iter():
            if child.tag.endswith("}v") or child.tag == "v":
                try:
                    slot = int(child.text or "")
                    if 0 <= slot < len(rel_slots):
                        return rel_slots[slot]
                except ValueError:
                    continue
    return None


def _count_richdata_cells(archive: zipfile.ZipFile, sheet_path: str) -> int:
    root = ET.fromstring(archive.read(sheet_path))
    return sum(1 for cell in _iter_local_name(root, "c") if "vm" in cell.attrib)


def _iter_local_name(root: ET.Element, local_name: str) -> list[ET.Element]:
    suffix = f"}}{local_name}"
    return [el for el in root.iter() if el.tag.endswith(suffix) or el.tag == local_name]
