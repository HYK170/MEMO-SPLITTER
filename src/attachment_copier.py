from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.filename_builder import unique_dest_path
from src.hyperlink_resolver import resolve_local_path
from src.image_handler import get_image_row
from src.xlsx_hyperlink_index import XlsxHyperlinkIndex


@dataclass
class AttachmentCopyResult:
    copied: list[str] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)


def collect_row_hyperlink_targets(
    ws: Worksheet,
    row: int,
    image_index: XlsxHyperlinkIndex | None = None,
) -> list[str]:
    targets: list[str] = []
    seen: set[str] = set()

    min_col = ws.min_column or 1
    max_col = ws.max_column or 1
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        hyperlink = cell.hyperlink
        if hyperlink is None:
            continue
        target = hyperlink.target or hyperlink.location
        if target and target not in seen:
            seen.add(target)
            targets.append(target)

    for image in getattr(ws, "_images", []):
        if get_image_row(image) != row:
            continue
        target = _get_image_hyperlink_target(image)
        if target and target not in seen:
            seen.add(target)
            targets.append(target)

    if image_index is not None:
        for target in image_index.targets_for_row(row):
            if target not in seen:
                seen.add(target)
                targets.append(target)

    return targets


def copy_attachments_for_row(
    ws: Worksheet,
    row: int,
    input_xlsx: Path,
    dest_folder: Path,
    image_index: XlsxHyperlinkIndex | None = None,
) -> AttachmentCopyResult:
    base_dir = input_xlsx.parent
    result = AttachmentCopyResult()
    targets = collect_row_hyperlink_targets(ws, row, image_index)

    for target in targets:
        local_path = resolve_local_path(target, base_dir)
        if local_path is None:
            if target.startswith("#"):
                result.skipped.append(f"내부 링크 스킵: {target}")
            elif target.lower().startswith(("http://", "https://", "mailto:")):
                result.skipped.append(f"원격 링크 스킵: {target}")
            else:
                result.skipped.append(f"경로 해석 실패: {target}")
            continue

        if not local_path.is_file():
            result.skipped.append(f"파일 없음: {local_path}")
            continue

        dest_path = unique_dest_path(dest_folder, local_path.name)
        try:
            shutil.copy2(local_path, dest_path)
            result.copied.append(str(dest_path.name))
        except OSError as exc:
            result.skipped.append(f"복사 실패 ({local_path.name}): {exc}")

    return result


def _get_image_hyperlink_target(image) -> str | None:
    anchor = getattr(image, "anchor", None)
    if anchor is None:
        return None

    pic = getattr(anchor, "pic", None)
    if pic is None:
        return None

    nv_pic_pr = getattr(pic, "nvPicPr", None)
    if nv_pic_pr is None:
        return None

    c_nv_pr = getattr(nv_pic_pr, "cNvPr", None)
    if c_nv_pr is None:
        return None

    hlink = getattr(c_nv_pr, "hlinkClick", None) or getattr(c_nv_pr, "hlinkHover", None)
    if hlink is None:
        return None

    for attr in ("target", "action"):
        value = getattr(hlink, attr, None)
        if value:
            return str(value)
    return None


def list_sheet_names(xlsx_path: Path) -> list[str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
